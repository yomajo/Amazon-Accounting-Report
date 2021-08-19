from openpyxl.utils import get_column_letter
from datetime import datetime
import platform
import logging
import json
import sys
import os

# GLOBAL VARIABLES
VBA_ERROR_ALERT = 'ERROR_CALL_DADDY'


def get_level_up_abspath(absdir_path):
    '''returns directory absolute path one level up from passed abs path'''
    return os.path.dirname(absdir_path)


def get_output_dir(client_file=True):
    '''returns target dir for output files depending on execution type (.exe/.py) and file type (client/systemic)'''
    # pyinstaller sets 'frozen' attr to sys module when compiling
    if getattr(sys, 'frozen', False):
        curr_folder = os.path.dirname(sys.executable)
    else:
        curr_folder = os.path.dirname(os.path.abspath(__file__))
    return get_level_up_abspath(curr_folder) if client_file else curr_folder

def file_to_binary(abs_fpath:str):
    '''returns binary data for file'''
    try:
        with open(abs_fpath, 'rb') as f:
            bfile = f.read()
        return bfile
    except FileNotFoundError as e:
        print(f'file_to_binary func got arg: {abs_fpath}; resulting in error: {e}')
        return None

def recreate_txt_file(abs_fpath:str, binary_data):
    '''outputs a file from given binary data'''
    try:
        with open(abs_fpath, 'wb') as f:
            f.write(binary_data)
    except TypeError:
        print(f'Expected binary when writing contents to file {abs_fpath}')

def is_windows_machine() -> bool:
    '''returns True if machine executing the code is Windows based'''
    machine_os = platform.system()
    return True if machine_os == 'Windows' else False

def orders_column_to_file(orders:list, dict_key:str):
    '''exports a column values of each orders list item for passed dict_key'''
    try:
        export_data = [order[dict_key] for order in orders]
        with open(f'export {dict_key}.txt', 'w', encoding='utf-8', newline='\n') as f:
            f.writelines('\n'.join(export_data))
        print(f'Data exported to: {os.path.dirname(os.path.abspath(__file__))} folder')
    except KeyError:
        print(f'Provided {dict_key} does not exist in passed orders list of dicts')

def alert_vba_date_count(filter_date, orders_count):
    '''Passing two variables for VBA to display for user in message box'''
    print(f'FILTER_DATE_USED: {filter_date}')
    print(f'SKIPPING_ORDERS_COUNT: {orders_count}')

def get_datetime_obj(date_str):
    '''returns tz-naive datetime obj from date string. Designed to work with str format: 2020-04-16T10:07:16+00:00'''
    try:
        return datetime.fromisoformat(date_str).replace(tzinfo=None)
    except ValueError:
        # Attempt to handle wrong/new date format here
        logging.warning(f'Change in format detected! Previous format: 2020-04-16T10:07:16+00:00. Current: {date_str} Attempting to parse string...')
        try:
            date_str_split = date_str.split('T')[0]
            return datetime.fromisoformat(date_str_split)
        except ValueError:
            logging.critical(f'Unable to create datetime from date string: {date_str}. Terminating.')
            print(VBA_ERROR_ALERT)
            sys.exit()

def simplify_date(date_str : str) -> str:
    '''returns a simplified date format: YYYY-MM-DD from rawformat 2020-04-16T06:53:44+00:00'''
    try:
        date = get_datetime_obj(date_str).date()
        return date.strftime('%Y-%m-%d')
    except ValueError:
        logging.warning(f'Unable to return simplified version of date: {date_str}. Returning raw format instead')
        return date_str

def col_to_letter(col : int, zero_indexed=True) -> str:
    '''returns column letter from worksheet column index'''
    if zero_indexed:
        col += 1
    return get_column_letter(col)

def get_last_used_row_col(ws:object):
    '''returns dictionary containing max_row and max_col as integers - last used row and column in passed openpyxl worksheet'''
    row = ws.max_row
    while row > 0:
        cells = ws[row]
        if all([cell.value is None for cell in cells]):
            row -= 1
        else:
            break
    if row == 0:
        return {'max_row' : 0, 'max_col' : 0}

    column = ws.max_column
    while column > 0:
        cells = next(ws.iter_cols(min_col=column, max_col=column, max_row=row))
        if all([cell.value is None for cell in cells]):
            column -= 1
        else:
            break
    return {'max_row' : row, 'max_col' : column}

def export_json_data(dataobj : dict, json_path : str ='export.json'):
    '''exports dataobj in json format'''
    with open(json_path, 'w') as f:
        json.dump(dataobj, f, indent=4)

def get_EU_countries_from_txt(txt_abspath:str) -> list:
    '''reads countries ISO codes listed each on new line and returns list of EU member countries'''
    countries = []
    try:
        with open(txt_abspath, 'r') as f:
            for line in f.readlines():
                countries.append(line.strip())
        return countries
    except Exception as e:
        logging.critical(f'Error reading EU countries from txt file: {txt_abspath}. Err: {e}. Alerting VBA, terminating immediately.')
        print(VBA_ERROR_ALERT)
        sys.exit()

def get_order_tax(order:dict) -> float:
    '''returns tax of order dict as float'''
    tax = float(order['item-tax'])
    return round(tax, 2)

if __name__ == "__main__":
    pass