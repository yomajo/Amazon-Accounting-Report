import copy
from collections import defaultdict
import openpyxl
from constants import TEMPLATE_SHEET_MAPPING, EU_SUMMARY_HEADERS, COM_SUMMARY_HEADERS
from accounting_utils import simplify_date, col_to_letter, get_last_used_row_col


# GLOBAL VARIABLES
SUMMARY_SHEET_NAME = 'Summary'
BOLD_STYLE = openpyxl.styles.Font(bold=True, name='Calibri')
BACKGROUND_COLOR_STYLE = openpyxl.styles.PatternFill(fgColor='D6DEFF',fill_type='solid')
THIN_BORDER = openpyxl.styles.Side(border_style='thin')
SHEET_HEADERS = list(TEMPLATE_SHEET_MAPPING.keys())
REPORT_START_ROW = 1
REPORT_START_COL = 1


class AmazonOrdersReport():
    '''Generic report class for both EU and COM reports
    Used to be AmazonEUOrdersReport. Since 2021-08 for easier differentiation separated to generic + inheritance for both report classes'''
    
    def __init__(self, export_obj: dict, eu_countries: list, sales_channel: str, proxy_keys: dict):
        self.eu_countries = eu_countries
        self.sales_channel = sales_channel
        self.proxy_keys = proxy_keys
        self.export_obj = self._clean_incoming_data(export_obj)
        self._get_report_objs()

    def _clean_incoming_data(self, export_obj:dict) -> dict:
        '''returns cleaned data in same structure as class input obj'''
        for region, currency in self._unpack_export_obj(export_obj):
            export_obj[region][currency] = self._correct_orders_numbers_dates(export_obj[region][currency])
        return export_obj

    @staticmethod
    def _unpack_export_obj(export_obj:dict):
        '''generator yielding self.export_obj dict KEY1 (region) and nested KEY2 (currency) at a time'''
        for region in export_obj:
            for currency in export_obj[region]:
                yield region, currency    
    
    def _correct_orders_numbers_dates(self, orders: list) -> list:
        '''date and numbers data cleaning for report:
        - original date format (2020-04-16T10:07:16+00:00) simplified to YYYY-MM-DD
        - all numbers of interest converted to floats'''
        for order in orders:
            # Simplifying order dates:
            order[self.proxy_keys['purchase-date']] = simplify_date(order[self.proxy_keys['purchase-date']], self.sales_channel)
            order[self.proxy_keys['payments-date']] = simplify_date(order[self.proxy_keys['payments-date']], self.sales_channel)
            # Converting numbers stored as strings to floats:
            order[self.proxy_keys['item-price']] = float(order[self.proxy_keys['item-price']])
            order[self.proxy_keys['item-tax']] = float(order[self.proxy_keys['item-tax']])
            order[self.proxy_keys['shipping-price']] = float(order[self.proxy_keys['shipping-price']])
            order[self.proxy_keys['shipping-tax']] = float(order[self.proxy_keys['shipping-tax']])
        return orders

    def _get_report_objs(self):
        '''prepares cls variables for excel report workbook filling'''
        self.segments_orders_obj = self._get_segments_orders_obj(self.export_obj)
        self.summary_table_obj = self._get_summary_table_obj(self.export_obj)
        self.summary_taxes_obj = self._get_summary_taxes_obj()

    def _get_segments_orders_obj(self, export_obj:dict) -> dict:
        '''returns dict of dicts for each segment (sheets) and corresponding list of orders (written to separate sheets)
        returned object: {'EU EUR' : [order1, order2, ...], 'NON-EU GBP' : [order11, order 22...], ...}'''
        segments_orders = {}
        for region, currency in self._unpack_export_obj(export_obj):
            segments_orders[f'{region} {currency}'] = export_obj[region][currency]
        return segments_orders

    def _get_summary_table_obj(self, export_obj: dict) -> dict:
        '''Returns nested object based on 1. currency 2. order dates. Example:
        {'EUR':{'date1':[order1, order2...], 'date2':[order3, order4...], ...},
        'GBP':{'date1':[order1, order2...], 'date2':[order1, order2...], ...}, ...}

        NOTE: output is region agnostic. Regions get mixed up'''
        summary_table_obj = defaultdict(dict)
        currency_based = defaultdict(list)
        # 1st loop forms currency (key) based obj. Output: {'currency1':[order1, order2, ...], 'currency2':[order2, order3...],...}
        for region, currency in self._unpack_export_obj(export_obj):
            currency_based[currency] = currency_based[currency] + export_obj[region][currency]
        # 2nd loop - forms {'currency':{'date1':[order1, order2, ...], 'date2':[order1, order2, ...], ...}, 'currency2':{...}}
        for currency in currency_based:
            summary_table_obj[currency] = self._get_payment_date_obj(currency_based[currency])
        return summary_table_obj

    def _get_payment_date_obj(self, orders: list) -> dict:
        '''forms a dictionary based on dates in list of order dicts. Returns payment dates as keys and orders list as values.
        Output format: {{'YYYY-MM-D1':[order1, order2, ...]},
                        {'YYYY-MM-D2':[order1, order2, ...]}, ...}'''
        payment_date_orders = defaultdict(list)
        for order in orders:
            payment_date_orders[order[self.proxy_keys['payments-date']]].append(order)
        return payment_date_orders

    def _get_summary_taxes_obj(self):
        '''Returns currency and date based calculated taxes for UK orders (item-tax + shipping)
        
        self.summary_table_obj:
        {'EUR':{'date1':[order1, order2...], 'date2':[order3, order4...], ...},
        'GBP':{'date1':[order1, order2...], 'date2':[order1, order2...], ...}, ...}

        function returns:
        {'EUR':{'date1':calculated_taxes1, 'date2':calculated_taxes2, ...},
        'GBP':{'date1':calculated_taxes3, 'date2':calculated_taxes4, ...}, ...}'''
        taxes_obj = copy.deepcopy(self.summary_table_obj)
        for currency, date_orders_dict in self.summary_table_obj.items():
            for date, date_orders in date_orders_dict.items():
                taxes_obj[currency][date] = self._calc_orders_taxes(date_orders, 'GB')
        return taxes_obj

    def _calc_orders_taxes(self, orders:list, country_code:str) -> float:
        '''returns sum of (item-tax + shipping-tax) for each order inside list if order[ship-country] = country_code'''
        taxes = 0
        for order in orders:
            if order[self.proxy_keys['ship-country']] == country_code:
                taxes += order[self.proxy_keys['item-tax']]
                taxes += order[self.proxy_keys['shipping-tax']]
        return round(taxes, 2)


    def _data_to_sheet(self, ws_name: str, orders_data: list):
        '''creates new ws_name sheet and fills it with orders_data argument data'''
        self._create_sheet(ws_name)
        active_ws = self.wb[ws_name]
        active_ws.freeze_panes = active_ws['A2']
        self._fill_sheet(active_ws, orders_data)
        self._adjust_col_widths(active_ws, self.col_widths)

    def _create_sheet(self, ws_name: str):
        '''creates new sheet obj with name ws_name'''
        self.wb.create_sheet(title=ws_name)
        self.col_widths = {}

    def _fill_sheet(self, active_ws, orders_data:list):
        '''writes headers, and corresponding orders data to active_ws, resizes columns'''
        self._write_headers(active_ws)
        self._write_sheet_orders(active_ws, orders_data)

    def _write_headers(self, ws: object):
        '''writes header row to segment sheet'''
        for col, header in enumerate(SHEET_HEADERS):
            self._update_col_widths(col, header)
            ws.cell(1, col + 1).value = header

    def _update_col_widths(self, col: int, cell_value: str, zero_indexed=True):
        '''runs on each cell. Forms a dictionary {'A':30, 'B':15...} for max column widths in worksheet (width as length of max cell)'''
        col_letter = col_to_letter(col, zero_indexed=zero_indexed)
        if col_letter in self.col_widths:
            # check for length, update if current cell length exceeds current entry for column
            if len(cell_value) > self.col_widths[col_letter]:
                self.col_widths[col_letter] = len(cell_value)
        else:
            self.col_widths[col_letter] = len(cell_value)

    def _write_sheet_orders(self, ws, orders_data: list):
        '''writes orders_data to segment sheet'''
        for row, col in self.range_generator(orders_data, SHEET_HEADERS):
            order_dict = orders_data[row]

            proxy_key = TEMPLATE_SHEET_MAPPING[SHEET_HEADERS[col]]
            # proxy value = order key
            proxy_value = self.proxy_keys[proxy_key]

            self._update_col_widths(col, str(order_dict[proxy_value]))
            # offsets due to excel vs python numbering  + headers in row 1
            ws.cell(row + 2, col + 1).value = order_dict[proxy_value]

    @staticmethod
    def range_generator(orders_data, headers):
        for row, _ in enumerate(orders_data):
            for col, _ in enumerate(headers):
                yield row, col

    def _adjust_col_widths(self, ws, col_widths: dict, summary=False):
        '''iterates over {'A':30, 'B':40, 'C':35...} dict to resize worksheets' column widths. Summary ws wider columns with summary=True'''
        factor = 1.3 if summary else 1.05
        for col_letter in col_widths:
            adjusted_width = ((col_widths[col_letter] + 2) * factor)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def fill_format_summary(self):
        '''Forms a summary sheet report unpacks self.summary_table_obj to dynamic height table,
        change insert point of table with:
        REPORT_START_ROW, REPORT_START_COL'''
        self.s_ws = self.wb[SUMMARY_SHEET_NAME]
        self.col_widths = {}   
        self.eu_countries_header_cols = {}
        self.row_cursor = REPORT_START_ROW
        self._add_summary_headers()
        # Add data for each currency:
        for currency, date_objs in self.summary_table_obj.items():
            self._apply_horizontal_line(self.row_cursor)
            self.s_ws.cell(self.row_cursor, REPORT_START_COL).value = currency
            self.s_ws.cell(self.row_cursor, REPORT_START_COL).font = BOLD_STYLE
            # Writing data to rest of columns:
            for date, date_orders in date_objs.items():
                self.s_ws.cell(self.row_cursor, REPORT_START_COL + 1).value = date
                self._fill_format_date_data(date_orders)
                uk_tax = self.summary_taxes_obj[currency][date]
                self.s_ws.cell(self.row_cursor, REPORT_START_COL + 7).value = uk_tax
                self.row_cursor += 1
        self._color_table_headers()
        self._adjust_col_widths(self.s_ws, self.col_widths, summary=True)

    def _add_summary_headers(self):
        '''writes fixed headers in summary sheet, freeze panes'''
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 4).value = 'Daily Breakdown'
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 4).font = BOLD_STYLE
        self.s_ws.freeze_panes = self.s_ws[f'C{REPORT_START_ROW + 2}']
        self.row_cursor += 1
        for idx, header in enumerate(EU_SUMMARY_HEADERS):
            self.s_ws.cell(self.row_cursor, REPORT_START_COL + idx).value = header
            self._update_col_widths(REPORT_START_COL + idx, header, zero_indexed=False)
            self.s_ws.cell(self.row_cursor, REPORT_START_COL + idx).font = BOLD_STYLE
        self.row_cursor += 1

    def _color_table_headers(self):
        '''Colors range defined in generator in summary sheet'''
        # max col used should be len of keys in self.col_widths
        max_col = len(self.col_widths.keys())
        for row, col in self._header_cells_generator(max_col):
            self.s_ws.cell(row, col).fill = BACKGROUND_COLOR_STYLE
    
    @staticmethod
    def _header_cells_generator(max_col: int):
        '''generator for daily breakdown headers coloring, yields row, col'''
        for row in [REPORT_START_ROW, REPORT_START_ROW + 1]:
            for col in range(REPORT_START_COL, max_col):
                yield row, col

    def _apply_horizontal_line(self, row: int):
        '''adds horizonal line through 100 cols (c=1 case) in summary sheet at argument row top'''
        for col in range(REPORT_START_COL, REPORT_START_COL + 99):
            self.s_ws.cell(row, col).border = openpyxl.styles.Border(top=THIN_BORDER)
    

    def _fill_format_date_data(self, date_orders: list):
        '''calculates neccessary data and fills, formats data in summary sheet in single row'''
        # Data does not update column widths, only headers. If data formats, scope were to change, function shall be updated 
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 2).value = self._get_segment_total(date_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 2).font = BOLD_STYLE
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 2).number_format = '#,##0.00'
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 3).value = len(date_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 3).font = BOLD_STYLE
        # Separating regions, filling data:
        eu_orders, non_eu_orders = self._split_by_region(date_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 4).value = self._get_segment_total(non_eu_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 4).number_format = '#,##0.00'
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 5).value = len(non_eu_orders)
        self._fill_summary_country_columns(eu_orders)

    def _split_by_region(self, orders: list):
        '''splits provided list of orders into two lists based on order['ship-country'] (using proxy keys) EU membership:
        1. eu orders
        2. non-eu orders'''    
        eu_orders, non_eu_orders = [], []
        for order in orders:
            if order[self.proxy_keys['ship-country']] in self.eu_countries:
                eu_orders.append(order)
            else:
                non_eu_orders.append(order)
        return eu_orders, non_eu_orders

    def _get_segment_total(self, orders: list) -> float:
        '''Returns a sum of all orders' item-price + shipping-price in list of order dicts'''
        total = 0
        for order in orders:
            total += order[self.proxy_keys['item-price']] + order[self.proxy_keys['shipping-price']]
        return round(total, 2)

    def _fill_summary_country_columns(self, eu_orders: list):
        '''fills individual eu countries data to separate columns'''
        countries_orders = self._get_country_based_dict(eu_orders)
        # Iterate countries, identify target/new column
        for country, country_orders in countries_orders.items():
            if country in self.eu_countries_header_cols.keys():
                ref_col = self.eu_countries_header_cols[country]
            else:
                last_used_row_col = get_last_used_row_col(self.s_ws)
                ref_col = last_used_row_col['max_col'] + 1
                # Enter data for existing self.row_cursor and new_col
                self._enter_new_country_header(country, ref_col)  

            # Add corresponding data in added/existing ref_col
            self._enter_format_country_date_data(country_orders, country, ref_col)

    def _get_country_based_dict(self, orders: list) -> dict:
        '''Splits list of orders to country based dict
        Returns:    {'DE': [order1, order2, ...], 'IT': [order1, order2, ...], ...}'''
        country_based_dict = defaultdict(list)
        for order in orders:
            country_code = order[self.proxy_keys['ship-country']]
            country_based_dict[country_code].append(order)
        return country_based_dict
    
    def _enter_new_country_header(self, country: str, ref_col: int):
        '''enter new country header col values, adjust col widths'''
        self.__enter_header_bold_update_col_widths(REPORT_START_ROW + 1, ref_col, f'{country} Sum')
        self.__enter_header_bold_update_col_widths(REPORT_START_ROW + 1, ref_col + 1, f'{country} #')
        self.__enter_header_bold_update_col_widths(REPORT_START_ROW + 1, ref_col + 2, f'{country} Taxes')
        # # gap column between countries
        self.__enter_header_bold_update_col_widths(REPORT_START_ROW + 1, ref_col + 3, ' ')
        # Update cls col reference dict
        self.eu_countries_header_cols[country] = ref_col
    
    def __enter_header_bold_update_col_widths(self, row: int, col: int, header: str):
        '''enters header value, bolds it, updates col widths dict'''
        self.s_ws.cell(row, col).value = header
        self._update_col_widths(col, header, zero_indexed=False)
        self.s_ws.cell(row, col).font = BOLD_STYLE

    def _enter_format_country_date_data(self, country_orders: list, country: str, ref_col: int):
        '''add total, count, taxes for currency>date>country orders at self.row_cursor, ref_col, formats number format'''
        self.s_ws.cell(self.row_cursor, ref_col).value = self._get_segment_total(country_orders)
        self.s_ws.cell(self.row_cursor, ref_col).number_format = '#,##0.00'
        self.s_ws.cell(self.row_cursor, ref_col + 1).value = len(country_orders)
        self.s_ws.cell(self.row_cursor, ref_col + 2).value = self._calc_orders_taxes(country_orders, country)


    def export(self, wb_name: str):
        '''Creates workbook, and exports class objects: segments_orders_obj and summary_table_obj to
        segment worksheets and creates report summary sheet, saves new workbook'''
        self.wb = openpyxl.Workbook()
        ws = self.wb.active
        ws.title = SUMMARY_SHEET_NAME
        for segment, segment_orders in self.segments_orders_obj.items():
            self._data_to_sheet(segment, segment_orders)
        self.fill_format_summary()
        self.wb.save(wb_name)
        self.wb.close()


    
    # ------------------ REVIEWED ABOVE ------------------------

class AmazonEUOrdersReport(AmazonOrdersReport):
    '''Intended for use of orders sold through AmazonEU / Amazon Warehouse sales channels
    accepts export data dictionary and output file path as arguments, creates individual
    sheets for region & currency based order segregation, creates formatted xlsx report file.
    Error handling is present outside of this class.

    Expected input obj format: 
        {eu_orders: {currency1: [order1, order2, order...], currency2: [order1, order2, order...], ...},
        non_eu_orders: {currency1: [order1, order2, order...], currency2: [order1, order2, order...] ...}}
    
    eu_countries: list of eu member countries as ['EE', 'LV', 'LT', 'FI', ...]
    
    Main method: export() - creates individual sheets, pushes selected data from corresponding orders;
    creates summary sheet, calculates regional / currency based totals'''

    def _split_by_region(self, orders:list):
        '''splits provided list of orders into two lists based on order['ship-country'] (using proxy keys) EU membership:
        1. eu orders
        2. non-eu orders
        NOTE: Specific to AMAZON EU report: orders with tax = 0 are attributed to non-EU'''    
        eu_orders, non_eu_orders = [], []
        for order in orders:
            if order[self.proxy_keys['ship-country']] in self.eu_countries:
                if order[self.proxy_keys['item-tax']] == 0:
                    non_eu_orders.append(order)
                else:
                    eu_orders.append(order)
            else:
                non_eu_orders.append(order)
        return eu_orders, non_eu_orders


class AmazonCOMOrdersReport(AmazonOrdersReport):
    '''Intended for use of orders sold through AmazonCOM sales channel. Simplified report version
    based on (inherited from) AmazonOrdersReport class
    
    accepts export data dictionary and output file path as arguments, creates individual
    sheets for region & currency based order segregation, creates formatted xlsx report file.
    Error handling is present outside of this class.

    Expected input obj format:
        {eu_orders: {currency1: [order1, order2, order...], currency2: [order1, order2, order...], ...},
        non_eu_orders: {currency1: [order1, order2, order...], currency2: [order1, order2, order...] ...}}
    
    eu_countries: list of eu member countries as ['EE', 'LV', 'LT', 'FI', ...]

    Main method: export() - creates individual sheets, pushes selected data from corresponding orders;
    creates summary sheet, calculates regional / currency based totals'''

    def _get_summary_taxes_obj(self):
        '''Returns currency and date based calculated taxes for UK orders (item-tax + shipping)
        
        self.summary_table_obj:
        {'EUR':{'date1':[order1, order2...], 'date2':[order3, order4...], ...},
        'GBP':{'date1':[order1, order2...], 'date2':[order1, order2...], ...}, ...}

        function returns:
        {'EUR':{'date1':calculated_taxes1, 'date2':calculated_taxes2, ...},
        'GBP':{'date1':calculated_taxes3, 'date2':calculated_taxes4, ...}, ...}'''
        taxes_obj = copy.deepcopy(self.summary_table_obj)
        for currency, date_orders_dict in self.summary_table_obj.items():
            for date, date_orders in date_orders_dict.items():
                taxes_obj[currency][date] = self._calc_orders_taxes(date_orders)
        return taxes_obj

    def _calc_orders_taxes(self, orders:list) -> float:
        '''returns sum of (item-tax + shipping-tax) for each order inside orders list'''
        taxes = 0
        for order in orders:
            taxes += order[self.proxy_keys['item-tax']]
            taxes += order[self.proxy_keys['shipping-tax']]
        return round(taxes, 2)

    def fill_format_summary(self):
        '''Forms a summary sheet report unpacks self.summary_table_obj to dynamic height table,
        change insert point of table with:
        REPORT_START_ROW, REPORT_START_COL'''
        self.s_ws = self.wb[SUMMARY_SHEET_NAME]
        self.col_widths = {}   
        self.row_cursor = REPORT_START_ROW
        self._add_summary_headers()
        self._color_table_headers()
        # Add data for each currency:
        for currency, date_objs in self.summary_table_obj.items():
            self._apply_horizontal_line(self.row_cursor)
            self.s_ws.cell(self.row_cursor, REPORT_START_COL).value = currency
            self.s_ws.cell(self.row_cursor, REPORT_START_COL).font = BOLD_STYLE
            # Writing data to rest of columns:
            for date, date_orders in date_objs.items():
                self.s_ws.cell(self.row_cursor, REPORT_START_COL + 1).value = date
                self._fill_format_date_data(date_orders)
                taxes = self.summary_taxes_obj[currency][date]
                self.s_ws.cell(self.row_cursor, REPORT_START_COL + 9).value = taxes
                self.row_cursor += 1
        # Adjust column widths
        self._adjust_col_widths(self.s_ws, self.col_widths, summary=True)

    def _add_summary_headers(self):
        '''writes fixed headers in summary sheet, freeze pane'''
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 4).value = 'Daily Breakdown'
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 4).font = BOLD_STYLE
        self.s_ws.freeze_panes = self.s_ws[f'A{REPORT_START_ROW + 2}']
        self.row_cursor += 1
        for idx, header in enumerate(COM_SUMMARY_HEADERS):
            self.s_ws.cell(self.row_cursor, REPORT_START_COL + idx).value = header
            self._update_col_widths(REPORT_START_COL + idx, header, zero_indexed=False)
            self.s_ws.cell(self.row_cursor, REPORT_START_COL + idx).font = BOLD_STYLE
        self.row_cursor += 1
    
    def _color_table_headers(self):
        '''Colors range defined in generator in summary sheet'''
        for row, col in self._header_cells_generator():
            self.s_ws.cell(row, col).fill = BACKGROUND_COLOR_STYLE

    @staticmethod
    def _header_cells_generator():
        '''generator for daily breakdown headers coloring, yields row, col'''
        for row in [REPORT_START_ROW, REPORT_START_ROW + 1]:
            for col in range(REPORT_START_COL, len(COM_SUMMARY_HEADERS) + REPORT_START_COL):
                yield row, col

    def _apply_horizontal_line(self, row:int):
        '''adds horizonal line through A:J (c=1 case) in summary sheet at argument row top'''
        for col in range(REPORT_START_COL, len(COM_SUMMARY_HEADERS) + REPORT_START_COL):
            self.s_ws.cell(row, col).border = openpyxl.styles.Border(top=THIN_BORDER)    

    def _fill_format_date_data(self, date_orders:list):
        '''calculates neccessary data and fills, formats data in summary sheet in single row'''
        # Data does not update column widths, only headers. If data formats, scope were to change, function shall be updated 
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 2).value = self._get_segment_total(date_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 2).font = BOLD_STYLE
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 2).number_format = '#,##0.00'
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 3).value = len(date_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 3).font = BOLD_STYLE
        # Separating regions, filling data:
        eu_orders, non_eu_orders = self._split_by_region(date_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 4).value = self._get_segment_total(eu_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 4).number_format = '#,##0.00'
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 5).value = len(eu_orders)

        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 6).value = self._get_segment_total(non_eu_orders)
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 6).number_format = '#,##0.00'
        self.s_ws.cell(self.row_cursor, REPORT_START_COL + 7).value = len(non_eu_orders)


if __name__ == "__main__":
    pass