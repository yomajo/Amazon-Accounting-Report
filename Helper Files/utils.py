import platform
import logging
import shutil
import json
import sys
import csv
import os
from datetime import datetime
from openpyxl.utils import get_column_letter
import charset_normalizer
from constants import VBA_ERROR_ALERT


def get_level_up_abspath(absdir_path):
    '''returns directory absolute path one level up from passed abs path'''
    return os.path.dirname(absdir_path)

def get_output_dir(client_file=True):
    '''returns target dir for output files depending on execution type (.exe/.py) and file type (client/systemic)'''
    # pyinstaller sets 'frozen' attr to sys module when compiling
    if getattr(sys, 'frozen', False):
        curr_folder = os.path.dirname(sys.executable)
    else:
        curr_folder = os.path.dirname(os.path.abspath(__file__))
    return get_level_up_abspath(curr_folder) if client_file else curr_folder

def is_windows_machine() -> bool:
    '''returns True if machine executing the code is Windows based'''
    machine_os = platform.system()
    return True if machine_os == 'Windows' else False

def orders_column_to_file(orders:list, dict_key:str):
    '''exports a column values of each orders list item for passed dict_key'''
    try:
        export_data = [order[dict_key] for order in orders]
        with open(f'export {dict_key}.txt', 'w', encoding='utf-8', newline='\n') as f:
            f.writelines('\n'.join(export_data))
        print(f'Data exported to: {os.path.dirname(os.path.abspath(__file__))} folder')
    except KeyError:
        print(f'Provided {dict_key} does not exist in passed orders list of dicts')

def alert_vba_date_count(filter_date, orders_count):
    '''Passing two variables for VBA to display for user in message box'''
    print(f'FILTER_DATE_USED: {filter_date}')
    print(f'SKIPPING_ORDERS_COUNT: {orders_count}')

def get_datetime_obj(date_str: str, sales_channel: str):
    '''returns tz-naive datetime obj from date string. Designed to work with str format: 2020-04-16T10:07:16+00:00'''
    try:
        if sales_channel == 'Amazon Warehouse':
            return datetime.strptime(date_str, '%Y.%m.%d  %H:%M:%S')
        else:
            # AmazonCOM / AmazonEU
            return datetime.fromisoformat(date_str).replace(tzinfo=None)
    except ValueError:
        logging.critical(f'Change in date format at sales channel: {sales_channel}! Could not parse to datetime: {date_str}. Terminating...')
        print(VBA_ERROR_ALERT)
        sys.exit()

def simplify_date(date_str: str, sales_channel: str) -> str:
    '''returns a simplified date format: YYYY-MM-DD from rawformat 2020-04-16T06:53:44+00:00'''
    try:
        date = get_datetime_obj(date_str, sales_channel).date()
        return date.strftime('%Y-%m-%d')
    except ValueError:
        logging.warning(f'Unable to return simplified version of date: {date_str}. Returning raw format instead')
        return date_str

def col_to_letter(col : int, zero_indexed=True) -> str:
    '''returns column letter from worksheet column index'''
    if zero_indexed:
        col += 1
    return get_column_letter(col)

def get_last_used_row_col(ws:object):
    '''returns dictionary containing max_row and max_col as integers - last used row and column in passed openpyxl worksheet'''
    row = ws.max_row
    while row > 0:
        cells = ws[row]
        if all([cell.value is None for cell in cells]):
            row -= 1
        else:
            break
    if row == 0:
        return {'max_row' : 0, 'max_col' : 0}

    column = ws.max_column
    while column > 0:
        cells = next(ws.iter_cols(min_col=column, max_col=column, max_row=row))
        if all([cell.value is None for cell in cells]):
            column -= 1
        else:
            break
    return {'max_row' : row, 'max_col' : column}

def dump_to_json(export_obj, json_fname:str) -> str:
    '''exports export_obj to json file. Returns path to crated json'''
    output_dir = get_output_dir(client_file=False)
    json_path = os.path.join(output_dir, json_fname)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(export_obj, f, indent=4)
    return json_path

def read_json_to_obj(json_file_path:str):
    '''reads json file and returns python object'''
    with open(json_file_path, 'r', encoding='utf-8') as f:
        orders = json.load(f)
    return orders

def get_EU_countries_from_txt(txt_abspath:str) -> list:
    '''reads countries ISO codes listed each on new line and returns list of EU member countries'''
    countries = []
    try:
        with open(txt_abspath, 'r') as f:
            for line in f.readlines():
                countries.append(line.strip())
        return countries
    except Exception as e:
        logging.critical(f'Error reading EU countries from txt file: {txt_abspath}. Err: {e}. Alerting VBA, terminating immediately.')
        print(VBA_ERROR_ALERT)
        exit()

def get_order_tax(order: dict, proxy_keys: dict) -> float:
    '''returns tax of order dict as float'''
    tax = float(order[proxy_keys['item-tax']])
    return round(tax, 2)

def get_file_encoding_delimiter(fpath:str) -> tuple:
    '''returns tuple of file encoding and delimiter'''
    with open(fpath, mode='rb') as f_as_bytes:
        try:
            byte_contents = f_as_bytes.read()
            enc_data = charset_normalizer.detect(byte_contents)
            encoding = enc_data['encoding']
            logging.info(f'Detected file encoding: {encoding}')
        except Exception as e:
            logging.warning(f'charset err: {e} when figuring out file {os.path.basename(fpath)} encoding. Defaulting to utf-8')
            encoding = 'utf-8'

    with open(fpath, mode='r', encoding=encoding) as f_text:
        text_contents = f_text.read()
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(text_contents)
        delimiter = dialect.delimiter if not dialect.delimiter == ' ' else '\t'
    return encoding, delimiter

def delete_file(file_abspath:str):
    '''deletes file located in file_abspath'''
    try:
        os.remove(file_abspath)
    except FileNotFoundError:
        logging.warning(f'Tried deleting file: {file_abspath}, but apparently human has taken care of it first. (File not found)')
    except Exception as e:
        logging.warning(f'Unexpected err: {e} while flushing db old records, deleting file: {file_abspath}')

def create_src_file_backup(target_file_abs_path:str, backup_fname_prefix:str) -> str:
    '''returns abspath of created file backup'''
    src_files_folder = get_src_files_folder()
    _, backup_ext = os.path.splitext(target_file_abs_path)
    backup_abspath = get_backup_f_abspath(src_files_folder, backup_fname_prefix, backup_ext)
    shutil.copy(src=target_file_abs_path, dst=backup_abspath)
    logging.info(f'Backup created at: {backup_abspath}')
    return backup_abspath

def get_src_files_folder():
    output_dir = get_output_dir(client_file=False)
    target_dir = os.path.join(output_dir, 'src files')
    if not os.path.exists(target_dir):
        os.mkdir(target_dir)
        logging.debug(f'src files directory inside Helper files has been recreated: {target_dir}')
    return target_dir

def get_backup_f_abspath(src_files_folder:str, backup_fname_prefix:str, ext:str) -> str:
    '''returns abs path for backup file. fname format: backup_fname_prefix-YY-MM-DD-HH-MM.ext'''
    timestamp = datetime.now().strftime('%y-%m-%d %H-%M')
    backup_fname = f'{backup_fname_prefix} {timestamp}{ext}'
    return os.path.join(src_files_folder, backup_fname)


if __name__ == "__main__":
    pass